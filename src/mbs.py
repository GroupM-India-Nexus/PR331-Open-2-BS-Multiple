import os
import glob
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from src.applyborders import apply_borders  # Assuming applyborders.py is in the same directory
from src.extract_datetime import extract_start_end_time  # Assuming extract_datetime.py is in the same directory

# Define the file paths
input_dir = "C:/Users/Sachin.Saurav/OneDrive - insidemedia.net/Multiple_Documents"
output_dir = os.path.join(input_dir, "output") #input_dir  # You can change this if you want outputs elsewhere

# Automatically detect all Excel files (excluding temp files like ~$.xlsx)
excel_files = [
 f for f in glob.glob(os.path.join(input_dir, "*.xlsx"))
 if not os.path.basename(f).startswith("~$")
]

# Create output directory if it doesn't exist
os.makedirs(output_dir, exist_ok=True)

# def extract_start_end_time(program_str):
#     match = re.search(r'\d{2}\.\d{2}', program_str)
#     if match:
#         times = re.findall(r'\d{2}\.\d{2}', program_str)
#         if len(times) == 2:
#             return times[0], times[1]
#     return "07.00", "24.00"  # Default values if format is incorrect

def process_tv_commercial_data(input_file, output_file):
    #output_file = os.path.join(output_dir, "Output_File_1.xlsx")
    
    # Load input workbook
    workbook = load_workbook(input_file, data_only=True)
    sheet = workbook.active
    
    # Extract headers and data
    headers = [cell.value for cell in sheet[1]]
    data = [{headers[i]: cell for i, cell in enumerate(row)} for row in sheet.iter_rows(min_row=2, values_only=True)]
    
    last_row = sheet.max_row
    # Extract 'CAMPAIGN_NAME' values from H2 to H47
    campaign_names = [sheet[f'H{row}'].value for row in range(2, last_row + 1)]
    
    # Extract Total Spots For Respective Camapigns
    #p_values = [sheet[f'P{row}'].value for row in range(2, 48)]

    # Extract 'COMMERCIAL_NAME' values from I2 to I47
    commercial_names = [sheet[f'I{row}'].value for row in range(2, last_row + 1)]

    # Extract 'Dur' values from B2 to B47
    dur_values = [sheet[f'B{row}'].value for row in range(2, last_row + 1)]

    # Extract 'Budget' values from L2 to L47
    net_outlay_values = [round(sheet[f'L{row}'].value) if sheet[f'L{row}'].value is not None else None for row in range(2, last_row + 1)]
    
    #net_outlay_values = [round(sheet[f'L{row}'].value) for row in range(2, 48)]

    # Extract 'GRP'
    grp_values = [sheet[f'M{row}'].value for row in range(2, last_row + 1)]
    
    # Extract 'Brand Names' values from J2 to J47
    brand_name = [sheet[f'J{row}'].value for row in range(2, last_row + 1)]
    
    # Extract 'IB NO' values from K2 to K47
    ib_no = [sheet[f'K{row}'].value for row in range(2, last_row + 1)]
    
    # Extract "Start Date" values from C2 to C47
    start_date = [sheet[f'C{row}'].value for row in range(2, last_row + 1)]
    
    # Extract 'End_Date' values from D2 to D47
    en_date = [sheet[f'D{row}'].value for row in range(2, last_row + 1)] 
     
    
    # Extract spot data from Q3 to AC47 for Q20 to Q30 (row-wise)
    #spot_values = [
        #[sheet.cell(row=row, column=col).value for col in range(17, 30)]
        #for row in range(2, 48)
    #]
    
    # extract lang values from O2 to O47
    lang_values = [sheet[f'O{row}'].value for row in range(2, last_row + 1)]

    
     # Extract 'Net Outlay', 'GRP', and 'TVR' values from L2, M2, and N2 to L47, M47, N47
    net_outlay_values = [round(sheet[f'L{row}'].value) if sheet[f'L{row}'].value is not None else None for row in range(2, last_row + 1)]
    #net_outlay_values = [sheet[f'L{row}'].value for row in range(2, 48)]
    grp_values = [sheet[f'M{row}'].value for row in range(2, last_row + 1)]
    #tvr_values = [sheet[f'N{row}'].value for row in range(2, 48)]
    
    # Extract date range
    try:
        ro_start_raw = data[0]['RO Start Date'] if data else None
        if ro_start_raw not in (None, '', 'n/a'):
            if isinstance(ro_start_raw, datetime):
                ro_start_date = ro_start_raw
            else:
                ro_start_date = datetime(1899, 12, 30) + timedelta(days=float(ro_start_raw))
        else:
            ro_start_date = datetime(2024, 9, 18)

        ro_end_raw = data[0]['RO End Date'] if data else None
        if ro_end_raw not in (None, '', 'n/a'):
            if isinstance(ro_end_raw, datetime):
                end_date = ro_end_raw
            else:
                end_date = datetime(1899, 12, 30) + timedelta(days=float(ro_end_raw))
        else:
            end_date = datetime(2024, 9, 30)
    except Exception as e:
        print("Date conversion error:", e)
        ro_start_date, end_date = datetime(2024, 9, 18), datetime(2024, 9, 30)
    
    dates = [ro_start_date + timedelta(days=i) for i in range((end_date - ro_start_date).days + 1)]
    
    #date_format = '%d-%b-%y'
    #try:
        #ro_start_date = datetime.strptime(data[0]['RO Start Date'], date_format) if data and data[0]['RO Start Date'] else datetime(2024, 9, 18)
        #end_date = datetime.strptime(data[0]['RO End Date'], date_format) if data and data[0]['RO End Date'] else datetime(2024, 9, 30)
    #except:
        #ro_start_date, end_date = datetime(2024, 9, 18), datetime(2024, 9, 30)

    # Generate dates range
    #dates = [ro_start_date + timedelta(days=i) for i in range((end_date - ro_start_date).days + 1)]
    
    intermediate_values = []
    rounded_results = []

    for row in range(2, last_row + 1):
        m_val = sheet[f'M{row}'].value
        n_val = sheet[f'N{row}'].value
        dur_val = dur_values[row - 2]  # Use the pre-fetched dur value

    # Avoid division by zero or NoneType errors
        if m_val is not None and n_val not in (0, None):
            result = m_val / n_val * 10
        else:
            result = None

        intermediate_values.append(result)

    # Apply =ROUND(O2/B2, 0) logic
        if result is not None and dur_val not in (0, None):
            rounded_value = round(result / dur_val)
        else:
            rounded_value = None

        rounded_results.append(rounded_value)

    # Create output workbook
    out_workbook = Workbook()
    out_worksheet = out_workbook.active  
    
    col_widths = {}
    
    # Define the range (P1 to BI17)
    start_col, end_col = 16, 15 + len(campaign_names) -1  # Column P (16) to BI (60)
    start_row, end_row = 1, 21
    
    # Define the range (P1 to BI17) 
    # Column P is 16th and BI is 60th in numerical index
    for row in out_worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):  
        for cell in row:
        # Apply center alignment
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Track max text length for each column
            if cell.value:
                col_widths[cell.column_letter] = max(col_widths.get(cell.column_letter, 0), len(str(cell.value)))
    
    # Adjust column width dynamically based on content
    for col_letter, max_length in col_widths.items():
        out_worksheet.column_dimensions[col_letter].width = max_length + 2  # Adding extra padding 
                               
    # Adjust row height for better spacing
    for row in range(start_row, end_row + 1):
        out_worksheet.row_dimensions[row].height = 21
    
#     # Define border styles : thin
#     medium_border = Border(
#     left=Side(style='medium'),
#     right=Side(style='medium'),
#     top=Side(style='medium'),
#     bottom=Side(style='medium')
# )
    
# #     # Function to apply borders to a given range
#     def apply_borders(out_worksheet, start_row, start_col, end_row, end_col):
#         for row in range(start_row, end_row + 1):
#             for col in range(start_col, end_col + 1):
#                 out_worksheet.cell(row=row, column=col).border = medium_border
    
#     # Apply borders to specific cell range
    apply_borders(out_worksheet, 3, 11, 6, 13)  # K3:M6
    apply_borders(out_worksheet, 10, 11, 13, 13)  # K10:M13
    apply_borders(out_worksheet, 9, 2, 11, 5)  # B9:E1

    light_yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    bottom_align = Alignment(vertical="bottom")

    cell_values = {
        'B9':'FCT LEFT', 'B10': '%Allocation', 'B11': 'CPRP', 
        'E9': 'Pristine', 'E10': '%', 'E11': 'CL.CPRP'
    }

    for cell_ref, text in cell_values.items():
        cell = out_worksheet[cell_ref]
        cell.value = text
        cell.fill = light_yellow_fill
        cell.font = bold_font

    hds = [
    "Input File F to G", "Input File E", "Input File A", "Input File N", "L/(M/N)",
    "Input File E : RODP-Start", "Input File E : RODP-End", "hardcoded to 1", "Day From Date", "'='Allocated",
    "sum P20 to CN20", "sum P20*P1-CN20*CN1", "N-M", "Hardcoded To As Per Deal"
    ]
    
    wrap_text_align = Alignment(vertical="bottom", wrapText=True)
    
    for col, value in enumerate(hds, start=2):  # B=2, C=3, ..., O=15
        cell = out_worksheet.cell(row=22, column=col, value=value) # 18 to 21 changed
        cell.font = bold_font
        cell.alignment = wrap_text_align

    # Define headers
    headers = [
        'ProgramIndex', 'Date', 'ProgramName', 'ChannelName', 'Rating', 'ER',
        'Start Time', 'End Time', 'DayPart', 'Day', 'Available', 'Spot', 'Allocated', 'Unallocated', 'PT or NPT'
    ]

    column_O_headers = [
        'Dur', 'Campaign Name', 'Commercial Name', 'Budget', 'GRP', 'PT', 'NPT', 'IB ID',
        'Amount', 'Allocated GRP', 'CPRP', 'GRP Allocation %', 'Budget Available',
        'Brand Name', 'IB No', 'Start Date', 'End Date', 'Total Dur', 'Variance GRP', 'Total Spots', 'Lang' 
    ]

    # Apply header formatting (light orange color)
    light_orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)

    # Write headers at row 19
    for col_num, header in enumerate(headers, start=1):
        cell = out_worksheet.cell(row=23, column=col_num, value=header) # 19 to 23 changed
        cell.fill = light_orange_fill
        cell.font = bold_font

    # Write headers at row 19 with light orange color till 'BK'
    for col_num in range(16, 15 + len(campaign_names) - 1):  # Columns A to BK (total 63 columns)
        cell = out_worksheet.cell(row=23, column=col_num) # 19 to 22 changed
        cell.fill = light_orange_fill
        cell.font = bold_font    

    # Write column O headers from O1 to O17
    for row_num, header in enumerate(column_O_headers, start=1):
        cell = out_worksheet.cell(row=row_num, column=15, value=header)  # Column O is the 15th column
        if row_num <= 8 or row_num >= 16:
            cell.fill = light_orange_fill
        elif 9 <= row_num <= 15:
            cell.fill = yellow_fill
        cell.font = bold_font

    # Standard data
    channel_name = data[0].get('Channels', 'ZEE TV') if data else 'ZEE TV'
    rating = data[0].get('TVR', 0.36) if data else 0.36
    cost_per_10_sec = data[0].get('ER', '0') if data else '0'

    day_mapping = {0: 'Mon', 1: 'Tue', 2: 'Wed', 3: 'Thu', 4: 'Fri', 5: 'Sat', 6: 'Sun'}
    availability = {'Mon': 1825, 'Tue': 1770, 'Wed': 1800, 'Thu': 1700, 'Fri': 1700, 'Sat': 1745, 'Sun': 1745}
    spot_counts = {'Mon': 95, 'Tue': 91, 'Wed': 93, 'Thu': 88, 'Fri': 88, 'Sat': 91, 'Sun': 91}

    # Write data rows starting from row 23
    for row_idx, date in enumerate(dates):
        day_name = day_mapping[date.weekday()]
        available = availability.get(day_name, 1700)
        spot_count = spot_counts.get(day_name, 90)
        campaign_name = data[row_idx].get('Campaign Name', '') if row_idx < len(data) else ''
        program_name = data[row_idx].get('Program', 'RODP-07.00-24.00') if row_idx < len(data) else 'RODP-07.00-24.00'
        start_time, end_time = extract_start_end_time(program_name)

        row_data = [
            row_idx + 1, date.strftime('%d/%m/%y'), program_name, channel_name, rating, cost_per_10_sec,
            start_time, end_time, 1, day_name, available, spot_count, available, 0, 'As Per Deal'
        ]

        row_data.extend([campaign_name])

        out_worksheet.append(row_data)

    start_col = 16  # Column 'P'
    start_row = 24  # Row where writing starts
    num_days = len(dates)

    for idx, total in enumerate(rounded_results):
        if total is None:
            continue  # Skip if no value to distribute

        # Base distribution and extras
        base_value = total // num_days
        extras = total % num_days

        # Create distribution list
        distribution = [base_value + 1 if i < extras else base_value for i in range(num_days)]

        # Get column letter for the current campaign
        col_letter = get_column_letter(start_col + idx)
        
        # Write each day's value vertically down the column
        for row_offset, value in enumerate(distribution):
            out_worksheet[f"{col_letter}{start_row + row_offset}"] = value
    
    
    # Write extracted spot values from Q3 to AC47 into P20 to BI30 (row-wise)
    #for col_idx, values in enumerate(spot_values, start=16):
        #for row_offset, value in enumerate(values, start=24): #20 to 23 changed
            #out_worksheet.cell(row=row_offset, column=col_idx, value=value)      
            
            
    # Define the range to fill (P6 to BI6)
    start_col = 16  # Column P
    end_col = 15 + len(campaign_names) - 1    # Column BI
    row = 6         # Row 6
    
    # Apply the formula dynamically for each column
    for col_idx in range(start_col, end_col + 1):  
        col_letter = out_worksheet.cell(row=1, column=col_idx).column_letter  
        formula = f"=IFERROR((SUMPRODUCT(($O$24:$O$34=$O{row})*($E$24:$E$34)*({col_letter}$1/10)*({col_letter}$24:{col_letter}$34))/{col_letter}10), 0)"
  
        out_worksheet.cell(row=row, column=col_idx, value=formula).font = bold_font
        
    
    # Define the range to fill (P7 to BI7)
    start_col = 16 #column P
    end_col = 15 + len(campaign_names) - 1  # Column BI
    row = 7
    
    # Apply the formula dynamically for each column
    for col_idx in range(start_col, end_col + 1):  
        col_letter = out_worksheet.cell(row=1, column=col_idx).column_letter  
        formula = f"=IFERROR(SUMPRODUCT(($O$24:$O$34=$O{row})*($E$24:$E$34)*({col_letter}$1/10)*({col_letter}$24:{col_letter}$34)), 0)"
  
        out_worksheet.cell(row=row, column=col_idx, value=formula).font = bold_font            

    # Write 'Dur' values from B2:B47 into P1 to BI1
    for col_idx, value in enumerate(dur_values, start=16):
        out_worksheet.cell(row=1, column=col_idx, value=value).font = bold_font # may add bold_font = Font(bold=True)
         
    # Write Total spots copied from P2 to P47 from Input to P20 to BI20 in Output  
    
    col_idx = 16
    for value in rounded_results:
        out_worksheet.cell(row=20, column=col_idx, value=value).font = bold_font
        col_idx += 1

       
    #for col_idx, value in enumerate(p_values, start=16):
       # out_worksheet.cell(row=20, column=col_idx, value=value).font = bold_font
    
    # Write 'Brand Names' from InputFile.xlsx    
    for col_idx, value in enumerate(brand_name, start=16):
        out_worksheet.cell(row=14, column=col_idx, value=value).font = bold_font # may add bold_font = Font(bold=True)
        
            
    # Write 'IB_NO' from InputFile.xlsx
    for col_idx, value in enumerate(ib_no, start=16):
        out_worksheet.cell(row=15, column=col_idx, value=str(value)).font = bold_font # may add bold_font = Font(bold=True)
        
    # write 'Start_Date' from InputFile.xlsx
    for col_idx, value in enumerate(start_date, start=16):
        out_worksheet.cell(row=16, column=col_idx, value=value).font = bold_font # may add bold_font = Font(bold=True)
        
    # Write 'End_Date' from InputFile.xlsx
    for col_idx, value in enumerate(en_date, start=16):
        out_worksheet.cell(row=17, column=col_idx, value=value).font = bold_font # may add bold_font = Font(bold=True)
    
    for col_idx, value in enumerate(lang_values, start=16):
        out_worksheet.cell(row=21, column=col_idx, value=value).font = bold_font
                              
    # Calculate and populate values under 'Allocated' (M20 to M32)
    for row_idx in range(24, 55):  # Rows 20 to 32
        spot_values_row = [out_worksheet.cell(row=row_idx, column=col_idx).value or 0 for col_idx in range(16, 15 + len(campaign_names) - 1)]  # P to BI
        allocated_value = sum(
            (dur_value or 0) * (spot_value or 0)
            for dur_value, spot_value in zip(dur_values, spot_values_row)
        )

        #allocated_value = sum(dur_value * spot_value for dur_value, spot_value in zip(dur_values, spot_values_row))
        out_worksheet.cell(row=row_idx, column=13, value=allocated_value)  # Column M is the 13th column    

    # Copy 'Available' values from M20 to M32 into 'Allocated' values in K20 to K32
    for row_idx in range(24, 55):  # Rows 20 to 32
        available_value = out_worksheet.cell(row=row_idx, column=13).value  # Column M (13th column)
        out_worksheet.cell(row=row_idx, column=11, value=available_value)  # Column K (11th column)
        
    # Put 'Spots' in P22
    out_worksheet["P23"] = "Spots" 
    out_worksheet["P23"].alignment = Alignment(horizontal="center", vertical="center")    
    
    total_grp_col = 16 + len(campaign_names) # Dynamically calculate the end column
    total_cost_col = total_grp_col + 1

    
    # Write 'TOTAL GRP' in BJ19 and 'TOTAL COST' in BK19
    out_worksheet.cell(row=23, column=total_grp_col, value='TOTAL GRP').fill = light_orange_fill
    out_worksheet.cell(row=23, column=total_grp_col).font = bold_font
    out_worksheet.cell(row=23, column=total_cost_col, value='TOTAL COST').fill = light_orange_fill
    out_worksheet.cell(row=23, column=total_cost_col).font = bold_font

    
    for row in range(3, 7):  # K3 to M6
        for col in range(11, 14):  # K=11, M=13 (inclusive)
            out_worksheet.cell(row=row, column=col).fill = yellow_fill
            
    out_worksheet["K3"] = "MIN"
    out_worksheet["K3"].font = bold_font
    out_worksheet["M3"] = "MAX"
    out_worksheet["M3"].font = bold_font
    out_worksheet["K4"] = "=ROUND(MIN($P$4:$BI$4), 0)"
    out_worksheet["K4"].font = bold_font
    out_worksheet["M4"] = "=ROUND(MAX($P$4:$BI$4), 0)"
    out_worksheet["M4"].font = bold_font
    out_worksheet["K5"] = "=MIN($P$5:$BI$5)"
    out_worksheet["K5"].font = bold_font
    out_worksheet["M5"] = "=MAX($P$5:$BI$5)"
    out_worksheet["M5"].font = bold_font
    out_worksheet["K6"] = "=MIN($P$6:$BI$6)"
    out_worksheet["K6"].font = bold_font
    out_worksheet["M6"] = "=MAX($P$6:$BI$6)"
    out_worksheet["M6"].font = bold_font
    
    for row in range(10, 14):  # K3 to M6
        for col in range(11, 14):  # K=11, M=13 (inclusive)
            out_worksheet.cell(row=row, column=col).fill = yellow_fill     
    
    out_worksheet["K10"] = "MIN"
    out_worksheet["K10"].font = bold_font
    out_worksheet["M10"] = "MAX"
    out_worksheet["M10"].font = bold_font
    out_worksheet["K11"] = "=MIN($P$11:$BI$11)"
    out_worksheet["K11"].font = bold_font
    out_worksheet["M11"] = "=MAX($P$11:$BI$11)"
    out_worksheet["M11"].font = bold_font
    out_worksheet["K12"] = "=MIN($P$12:$BI$12)"
    out_worksheet["K12"].font = bold_font
    out_worksheet["M12"] = "=MAX($P$12:$BI$12)"
    out_worksheet["M12"].font = bold_font
    out_worksheet["K13"] = "=MIN($P$13:$BI$13)"
    out_worksheet["K13"].font = bold_font
    out_worksheet["M13"] = "=MAX($P$13:$BI$13)"
    out_worksheet["M13"].font = bold_font
    
    
    # Define the light purple fill color
    light_purple_fill = PatternFill(start_color="E6CCFF", end_color="E6CCFF", fill_type="solid")

# Apply the fill to the range P12:BI12
    for col_idx in range(16, end_col + 1):  # P is 16th column, BI is 62nd column
        cell = out_worksheet.cell(row=12, column=col_idx)
        cell.fill = light_purple_fill      
             
    # Calculate and add the values for FCT LEFT
    out_worksheet.cell(row=9, column=3, value="=SUM(N20:N9530)")  # BJ18
    out_worksheet.cell(row=9, column=3).font = bold_font
    
    # Calculate and populate TOTAL GRP using formula
    for row_idx in range(24, 55): 
        formula = f"=ROUND(E{row_idx}*K{row_idx}/10, 0)"    
        out_worksheet.cell(row=row_idx, column=total_grp_col, value=formula)

    # Calculate and populate TOTAL COST using formula
    for row_idx in range(24, 55):  # Rows 20 to 32 # 20 to 23 changed
        formula_cost = f"=ROUND(F{row_idx}*K{row_idx}/1000000, 0)"
        out_worksheet.cell(row=row_idx, column=total_cost_col, value=formula_cost)  # BK column (63rd)

    

    # Calculate the formula (BK18 / BJ18) * 100000 and store in BL18
    bl_col = total_cost_col + 1
    out_worksheet.cell(row=18, column=bl_col, value= f"=IFERROR(( {get_column_letter(total_cost_col)}18 / {get_column_letter(total_grp_col)}18 ) * 100000, 0)")
    out_worksheet.cell(row=18, column=bl_col).font = bold_font  # BL18

    # Apply light yellow color to BJ18, BK18, BL18
    out_worksheet.cell(row=18, column=total_grp_col).fill = light_yellow_fill  # BJ18
    out_worksheet.cell(row=18, column=total_grp_col).fill = light_yellow_fill  # BK18
    out_worksheet.cell(row=18, column=bl_col).fill = light_yellow_fill  # BL18

    
    
    # Calculate CostPer10Sec dynamically and store in F20 to F32
    #for row_idx in range(24, 36):                                 # 20 to 23 and 33 to 36 changed
        #net_outlay = net_outlay_values[row_idx - 20] or 0
        #grp = grp_values[row_idx - 20] or 1  # Avoid division by zero
        #tvr = tvr_values[row_idx - 20] or 1  # Avoid division by zero
        #cost_per_10_sec = net_outlay / (grp / tvr) if grp / tvr != 0 else 0
        #out_worksheet.cell(row=row_idx, column=6, value=cost_per_10_sec)  # Column F (6th)

     # Write extracted 'CAMPAIGN_NAME' values row-wise from P2 to BI2
    for col_idx, value in enumerate(campaign_names, start=16):  # P is 16th column, BI is 62nd column
        out_worksheet.cell(row=2, column=col_idx, value=value).font = bold_font
        

    # Write commercial names from I2:I47 into P3 to BI3
    for col_idx, value in enumerate(commercial_names, start=16):
        out_worksheet.cell(row=3, column=col_idx, value=value).font = bold_font

    # Copy Net Outlay values from L2 to L47 into P4 to BI4 row-wise
    for col_idx, value in enumerate(net_outlay_values, start=16):  # P is 16th column
        out_worksheet.cell(row=4, column=col_idx, value=value).font = bold_font

    # Copy GRP values
    for col_idx, value in enumerate(grp_values, start=16):
        out_worksheet.cell(row=5, column=col_idx, value=value).font = bold_font
            
    # Apply SUMPRODUCT formula in P9 to BI9
    for col_idx in range(16, end_col + 1):  # P is 16th column, BI is 62nd column
        col_letter = out_worksheet.cell(row=1, column=col_idx).column_letter
        formula = f"=ROUND(SUMPRODUCT({col_letter}24:{col_letter}9530,$F$24:$F$9530)*{col_letter}$1/10, 0)"
        out_worksheet.cell(row=9, column=col_idx, value=formula).font = bold_font
        
    # Compute and write SUMPRODUCT formula in P10 to BI10
    for col_idx in range(16, end_col + 1):  # P is 16th column, BI is 62nd column
        col_letter = out_worksheet.cell(row=1, column=col_idx).column_letter
        formula = f"=ROUND(SUMPRODUCT({col_letter}24:{col_letter}9530,$E$24:$E$9530)*{col_letter}$1/10, 0)"
        out_worksheet.cell(row=10, column=col_idx, value=formula).font = bold_font    

    for col_idx in range(16, end_col + 1):  # P is 16th column, BI is 62nd column
        col_letter = out_worksheet.cell(row=1, column=col_idx).column_letter
        formula = f'=IFERROR(ROUND({col_letter}9/{col_letter}10, 0), 0)'
        out_worksheet.cell(row=11, column=col_idx, value=formula).font = bold_font
        
    for col_idx in range(16, end_col + 1): #P is 16th column, BI is 62nd column
        col_letter = out_worksheet.cell(row=1, column=col_idx).column_letter
        formula = f"=IFERROR(ROUND({col_letter}10/{col_letter}5*100, 0), 0)"
        out_worksheet.cell(row=12, column=col_idx, value=formula).font = bold_font
        
    for col_idx in range(16, end_col + 1):
        col_letter = out_worksheet.cell(row=1, column=col_idx).column_letter
        formula = f"={col_letter}4-{col_letter}9"
        out_worksheet.cell(row=13, column=col_idx, value=formula).font = bold_font
        
    for col_idx in range(16, end_col + 1):  # P is 16th column, BI is 62nd column # For Total Dur in row 18
        col_letter = out_worksheet.cell(row=1, column=col_idx).column_letter
        formula = f"=18*{col_letter}1"
        out_worksheet.cell(row=18, column=col_idx, value=formula).font = bold_font
        
    for col_idx in range(16, end_col + 1):  # P is 16th column, BI is 62nd column # For Variance GRP in row 19
        col_letter = out_worksheet.cell(row=1, column=col_idx).column_letter
        formula = f"={col_letter}10-{col_letter}24"
        out_worksheet.cell(row=19, column=col_idx, value=formula).font = bold_font           
    
    
    # Calculate the column index for the "TOTAL GRP" column
    total_grp_col_index = 16 + len(campaign_names)    
    
    # Convert the column index to a column letter
    total_grp_col_letter = get_column_letter(total_grp_col_index)    

    # Add SUM formula for TOTAL GRP in BJ18
    out_worksheet.cell(row=18, column=total_grp_col, value= f"=SUM({total_grp_col_letter}24:{total_grp_col_letter}55)")  # BJ18
    out_worksheet.cell(row=18, column=total_grp_col).font = bold_font

    # Store the SUM formula for TOTAL COST in BK18
    out_worksheet.cell(row=18, column=total_cost_col, value= f"=SUM({get_column_letter(total_cost_col)}24:{get_column_letter(total_cost_col)}55)")
    out_worksheet.cell(row=18, column=total_cost_col).font = bold_font
    out_worksheet.cell(row=18, column=total_cost_col).fill = light_yellow_fill    
                
    # Insert SUM formula in BJ4
    out_worksheet[f"{total_grp_col_letter}4"] = f"=ROUND(SUM(P4:{get_column_letter(end_col)}4), 0)"
    out_worksheet[f"{total_grp_col_letter}4"].fill = light_yellow_fill
    
    # Insert SUM formula in BJ5
    out_worksheet[f"{total_grp_col_letter}5"] = f"=ROUND(SUM(P5:{get_column_letter(end_col)}5), 0)"
    out_worksheet[f"{total_grp_col_letter}5"].fill = light_yellow_fill
    
    # Insert SUM formula in BJ9
    out_worksheet[f"{total_grp_col_letter}9"] = f"=ROUND(SUM(P9:{get_column_letter(end_col)}9), 0)"
    out_worksheet[f"{total_grp_col_letter}9"].fill = light_yellow_fill
    
    # Insert SUM formula in BJ10
    out_worksheet[f"{total_grp_col_letter}10"] = f"=ROUND(SUM(P10:{get_column_letter(end_col)}10), 0)"
    out_worksheet[f"{total_grp_col_letter}10"].fill = light_yellow_fill
    
    # Calculate the column index for BL
    bl_col_index = total_grp_col_index + 2
    bl_col_letter = get_column_letter(bl_col_index)

    
    out_worksheet[f"{bl_col_letter}17"] = f"=ROUND(({total_grp_col_letter}18/{total_grp_col_letter}5)*100, 0)"
    out_worksheet[f"{bl_col_letter}17"].fill = light_yellow_fill
    
    out_worksheet["C10"] = out_worksheet[f"{bl_col_letter}17"].value
    out_worksheet["C10"].font = bold_font
    out_worksheet["C11"] = out_worksheet[f"{total_grp_col_letter}18"].value
    out_worksheet["C11"].font = bold_font
    
    # Save output file
    out_workbook.save(output_file)
    print(f"Data processed and saved to {output_file}")

if __name__ == "__main__":

 if not excel_files:
     print("No valid Excel files found in the folder.")
 else:
     for input_file in excel_files:
         # Generate a unique output filename for each input file
         base_name = os.path.splitext(os.path.basename(input_file))[0]
         output_file = os.path.join(
             output_dir,
             f"Output_{base_name}.xlsx"
         )
         print(f"Processing {input_file} -> {output_file}")
         process_tv_commercial_data(input_file, output_file)

    #process_tv_commercial_data(input_file, output_dir)
